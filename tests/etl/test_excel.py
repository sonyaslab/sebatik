"""Tes adaptor .xlsx -> bentuk sumber dataset database.

Workbook dibangun di memori supaya tes tidak bergantung pada `data/raw/`
yang tidak ter-commit.
"""

from __future__ import annotations

import io

import pytest
from openpyxl import Workbook

from src.etl.database import DatasetTidakValid, transformasi_sumber_database
from src.etl.excel import SHEET_WAJIB, baca_workbook

HEADER_MASTER = [
    "ID Indikator",
    "Kategori",
    "Kelompok / Pilar",
    "Arah Pembangunan",
    "Kode Indikator",
    "Nama Indikator (RPJPD Provinsi / dipakai Kaltara)",
    "Indikator Proxy?",
    "Definisi (RPJPD Provinsi)",
    "Rumus Perhitungan (RPJPD Provinsi)",
    "Interpretasi (RPJPD Provinsi)",
    "Sumber Data (RPJPD Provinsi)",
    "Frekuensi (RPJPD Provinsi)",
    "Status Metadata",
    "Perangkat Daerah Pengampu (Kaltara)",
    "Ketersediaan Data",
    "Periode Data",
    "Tahun Data Terakhir",
    "Catatan Kualitas Data",
]

HEADER_NILAI = [
    "ID Indikator",
    "Kategori",
    "Kelompok / Pilar",
    "Kode Indikator",
    "Nama Indikator (Kaltara)",
    "Jenis Nilai",
    "Tahun",
    "Nilai (Angka)",
    "Nilai (Teks Asli)",
    "Satuan/Catatan",
]


def _id_uji(nomor: int) -> str:
    """86 indikator uji: 10 ISV lalu 76 IUP, sepola dengan berkas nyata."""
    if nomor <= 10:
        return f"ISV-{nomor:03d}"
    return f"IUP-{nomor - 10:03d}"


def _workbook_uji(jumlah: int = 86, nilai: list[list] | None = None, sheet: tuple[str, ...] = SHEET_WAJIB) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet[0]
    ws.append(HEADER_MASTER)
    for nomor in range(1, jumlah + 1):
        iid = _id_uji(nomor)
        ws.append(
            [
                iid,
                iid.split("-")[0],
                "Kelompok Uji",
                "Arah Uji",
                str(nomor),
                f"Indikator {iid}",
                "Tidak",
                f"Definisi {iid}",
                "a / b x 100",
                "Makin tinggi makin baik",
                "BPS",
                "Tahunan",
                "Lengkap",
                "BPS Kaltara",
                "Tersedia",
                "Tahunan",
                2025,
                None,
            ]
        )
    if len(sheet) > 1:
        ws2 = wb.create_sheet(sheet[1])
        ws2.append(HEADER_NILAI)
        for baris in nilai or []:
            ws2.append(baris)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def test_konversi_berhasil_dan_manifest_sesuai():
    isi = _workbook_uji(
        nilai=[
            ["ISV-001", "ISV", "Kelompok Uji", "1", "Indikator ISV-001", "Realisasi", 2021, 100.5, None, None],
            ["ISV-001", "ISV", "Kelompok Uji", "1", "Indikator ISV-001", "Target", 2025, 120.0, None, None],
        ]
    )
    sumber = baca_workbook(isi, "uji.xlsx")

    assert sumber["source"] == "uji.xlsx"
    assert set(sumber["sheets"]) == set(SHEET_WAJIB)

    dataset = transformasi_sumber_database(sumber)
    assert dataset["manifest"] == {"indikator": 86, "metadata_indikator": 86, "nilai_indikator": 2}


def test_sheet_hilang_ditolak_dengan_pesan_indonesia():
    isi = _workbook_uji(sheet=("Basis Data Indikator",))
    with pytest.raises(DatasetTidakValid, match="Data Target-Realisasi"):
        baca_workbook(isi, "uji.xlsx")


def test_byte_sampah_jadi_dataset_tidak_valid_bukan_traceback_openpyxl():
    with pytest.raises(DatasetTidakValid, match="bukan berkas .xlsx"):
        baca_workbook(b"ini jelas bukan excel", "palsu.xlsx")


def test_berkas_kosong_juga_ditolak():
    with pytest.raises(DatasetTidakValid):
        baca_workbook(b"", "kosong.xlsx")


def test_master_kurang_dari_86_baris_ditolak():
    sumber = baca_workbook(_workbook_uji(jumlah=85), "uji.xlsx")
    with pytest.raises(DatasetTidakValid, match="86 ID unik"):
        transformasi_sumber_database(sumber)


def test_sheet_nilai_kosong_diterima():
    """Keputusan desain §2: sheet nilai boleh parsial, bahkan kosong."""
    dataset = transformasi_sumber_database(baca_workbook(_workbook_uji(nilai=[]), "uji.xlsx"))
    assert dataset["manifest"]["nilai_indikator"] == 0
    assert dataset["manifest"]["indikator"] == 86


def test_sheet_nilai_parsial_diterima():
    isi = _workbook_uji(
        nilai=[["IUP-001", "IUP", "Kelompok Uji", "11", "Indikator IUP-001", "Realisasi", 2023, 7.0, None, None]]
    )
    dataset = transformasi_sumber_database(baca_workbook(isi, "uji.xlsx"))
    assert dataset["manifest"]["nilai_indikator"] == 1


def test_nilai_sel_non_primitif_dijadikan_teks():
    """Sel bertipe tanggal/kaya harus jadi str, sepola tools/import_classified_workbook.py."""
    from datetime import datetime

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_WAJIB[0]
    ws.append(["ID Indikator", "Periode Data"])
    ws.append(["ISV-001", datetime(2025, 1, 31)])
    ws2 = wb.create_sheet(SHEET_WAJIB[1])
    ws2.append(HEADER_NILAI)
    buffer = io.BytesIO()
    wb.save(buffer)

    sumber = baca_workbook(buffer.getvalue(), "uji.xlsx")
    sel = sumber["sheets"][SHEET_WAJIB[0]][1][1]
    assert isinstance(sel, str)
