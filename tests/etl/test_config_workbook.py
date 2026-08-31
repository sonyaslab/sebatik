"""Tes konfigurasi ETL.

Klaim yang diuji: menambah versi workbook baru cukup mengubah `workbook.yaml`,
tanpa menyentuh kode. Karena itu ada tes yang membaca workbook sintetis dengan
tata letak berbeda lewat konfigurasi buatan sendiri.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest
import yaml

from src.etl.config import KonfigurasiTidakValid, bawaan, muat
from src.etl.extract import baca_master, baca_nilai, lengkapi_pemilik
from src.etl.pipeline import JumlahIndikatorTidakSesuai, run

AKAR = Path(__file__).resolve().parents[2]
BERKAS_BAWAAN = AKAR / "src" / "etl" / "config" / "workbook.yaml"


# --- konfigurasi bawaan ----------------------------------------------------


def test_konfigurasi_bawaan_dapat_dimuat():
    konfigurasi = bawaan()
    assert konfigurasi.jumlah_indikator_diharapkan == 86
    assert konfigurasi.kategori == ("ISV", "IUP")
    assert konfigurasi.isv_nomor_maksimum == 10
    assert len(konfigurasi.nilai) == 4


def test_urutan_sumber_nilai_adalah_urutan_prioritas():
    """Sumber pertama menang; yang berikutnya hanya mengisi yang kosong."""
    assert [s.sheet for s in bawaan().nilai] == [
        "Rakor ISV IUP Kaltara 202607",
        "Rakor ISV IUP Kaltara 2026",
        "ISV IUP Kaltara 2026",
        "ISV IUP Kaltara",
    ]


def test_blok_menerjemahkan_rentang_kolom_menjadi_pasangan_tahun():
    sumber = bawaan().nilai[0]
    pasangan = sumber.blok[0].pasangan_kolom_tahun()
    assert pasangan[0] == (5, 2021)
    assert pasangan[-1] == (13, 2029)
    # Kolom horizon 2045 berdiri sendiri dan hanya berlaku untuk target.
    horizon = sumber.blok[1]
    assert horizon.pasangan_kolom_tahun() == [(14, 2045)]
    assert horizon.hanya_jenis == "target"


def test_baris_dibatasi_ukuran_sheet_sebenarnya():
    baris = bawaan().nilai[0].baris
    assert list(baris.rentang(10)) == list(range(3, 11))
    assert baris.rentang(1000)[-1] == 164


def test_baris_tanpa_akhir_membaca_sampai_habis():
    baris = bawaan().master.baris
    assert baris.akhir is None
    assert baris.rentang(50)[-1] == 50


def test_sheet_wajib_diturunkan_dari_konfigurasi():
    wajib = bawaan().sheet_wajib
    assert "form provinsi" in wajib
    assert len(wajib) == len(set(wajib))


# --- validasi --------------------------------------------------------------


def _tulis(tmp_path: Path, ubah) -> Path:
    data = yaml.safe_load(BERKAS_BAWAAN.read_text(encoding="utf-8"))
    ubah(data)
    berkas = tmp_path / "workbook.yaml"
    berkas.write_text(yaml.safe_dump(data, allow_unicode=True), encoding="utf-8")
    return berkas


def test_menolak_identitas_tak_dikenal(tmp_path):
    berkas = _tulis(tmp_path, lambda d: d["nilai"][0].__setitem__("identitas", "entah"))
    with pytest.raises(KonfigurasiTidakValid, match="identitas"):
        muat(berkas)


def test_menolak_blok_dari_kolom_tanpa_kolom_jenis(tmp_path):
    berkas = _tulis(tmp_path, lambda d: d["nilai"][0].pop("kolom_jenis"))
    with pytest.raises(KonfigurasiTidakValid, match="kolom_jenis"):
        muat(berkas)


def test_menolak_blok_tanpa_rentang_kolom(tmp_path):
    berkas = _tulis(tmp_path, lambda d: d["nilai"][1]["blok"][0].pop("kolom_akhir"))
    with pytest.raises(KonfigurasiTidakValid):
        muat(berkas)


def test_menolak_tanpa_sumber_nilai(tmp_path):
    berkas = _tulis(tmp_path, lambda d: d.__setitem__("nilai", []))
    with pytest.raises(KonfigurasiTidakValid, match="Minimal satu sumber"):
        muat(berkas)


def test_menolak_berkas_tidak_ada(tmp_path):
    with pytest.raises(KonfigurasiTidakValid, match="tidak ditemukan"):
        muat(tmp_path / "hilang.yaml")


# --- workbook versi lain ---------------------------------------------------


@pytest.fixture
def workbook_versi_lain(tmp_path: Path) -> Path:
    """Workbook dengan nama sheet, posisi kolom, dan tahun yang sama sekali beda."""
    wb = openpyxl.Workbook()
    master = wb.active
    master.title = "Master 2030"
    # Header sengaja diletakkan di baris 2 dan urutan kolomnya diacak.
    master.append(["catatan bebas"])
    master.append(["Nama Indikator", "Golongan", "Unit Pengampu", "Catatan"])
    master.append(["Angka Harapan Hidup", "ISV", "Dinas Kesehatan", None])
    master.append(["Panjang Jalan Mantap", "IUP", "Dinas PU", "Indikator proxy: Jalan provinsi"])

    nilai = wb.create_sheet("Angka 2030")
    nilai.append(["Golongan", "No", "Nilai 2030", "Nilai 2031"])
    nilai.append(["ISV", 1, 72.5, 73.0])
    nilai.append(["IUP", 11, 61.0, 62.5])

    pemilik = wb.create_sheet("Pengampu 2030")
    pemilik.append(["No", "Nama", "OPD"])
    pemilik.append([1, "Angka Harapan Hidup", "Dinas Kesehatan"])

    berkas = tmp_path / "workbook-2030.xlsx"
    wb.save(berkas)
    wb.close()
    return berkas


@pytest.fixture
def konfigurasi_versi_lain(tmp_path: Path) -> Path:
    data = {
        "versi": 2,
        "kategori": {"isv_nomor_maksimum": 10, "daftar": ["ISV", "IUP"]},
        "ekspektasi": {"jumlah_indikator": 2},
        "master": {
            "sheet": "Master 2030",
            "baris_header": 2,
            "baris": {"awal": 3, "akhir": None},
            "kolom": {
                "kategori": "Golongan",
                "nama_asli": "Nama Indikator",
                "catatan_teknis": "Catatan",
                "indikator_proxy": "Tidak Ada Kolom Ini",
            },
            "pic": {},
            "bawaan": {"status_metadata": "Tidak Tersedia"},
        },
        "pemilik": {
            "sheet": "Pengampu 2030",
            "identitas": "nomor_saja",
            "baris": {"awal": 2, "akhir": None},
            "kolom_nomor": 1,
            "kolom_nama": 2,
            "kolom_opd": 3,
        },
        "nilai": [
            {
                "sheet": "Angka 2030",
                "identitas": "kategori_nomor",
                "baris": {"awal": 2, "akhir": None},
                "kolom_kategori": 1,
                "kolom_nomor": 2,
                "blok": [
                    {
                        "jenis": "realisasi",
                        "kolom_awal": 3,
                        "kolom_akhir": 4,
                        "tahun_awal": 2030,
                    }
                ],
            }
        ],
        "audit": {"sheet": ["Master 2030"], "tahun_valid": [2020, 2050]},
    }
    berkas = tmp_path / "workbook-2030.yaml"
    berkas.write_text(yaml.safe_dump(data, allow_unicode=True), encoding="utf-8")
    return berkas


def test_workbook_versi_lain_terbaca_tanpa_mengubah_kode(workbook_versi_lain, konfigurasi_versi_lain):
    """Inti kriteria selesai etl.md: format baru = konfigurasi baru saja."""
    konfigurasi = muat(konfigurasi_versi_lain)
    wb = openpyxl.load_workbook(workbook_versi_lain, data_only=True)
    try:
        indikator, _ = baca_master(wb, konfigurasi)
        lengkapi_pemilik(wb, konfigurasi, indikator)
        nilai, statistik = baca_nilai(wb, konfigurasi)
    finally:
        wb.close()

    assert [x["id_indikator"] for x in indikator] == ["ISV-01", "IUP-01"]
    assert indikator[0]["nama_indikator"] == "Angka Harapan Hidup"
    assert indikator[0]["opd_penanggung_jawab"] == "Dinas Kesehatan"
    # Proxy tetap terbaca dari catatan meski kolom penandanya tidak ada.
    assert indikator[1]["is_proxy"] == 1
    assert indikator[1]["nama_proxy"] == "Jalan provinsi"

    assert statistik.berhasil == 4
    assert {(x["id_indikator"], x["tahun"], x["nilai"]) for x in nilai} == {
        ("ISV-01", 2030, 72.5),
        ("ISV-01", 2031, 73.0),
        ("IUP-01", 2030, 61.0),
        ("IUP-01", 2031, 62.5),
    }


def test_jumlah_indikator_diharapkan_berasal_dari_konfigurasi(tmp_path, workbook_versi_lain, konfigurasi_versi_lain):
    """Pesan galat menyebut nilai ekspektasi vs aktual, bukan angka 86 hardcode."""
    data = yaml.safe_load(konfigurasi_versi_lain.read_text(encoding="utf-8"))
    data["ekspektasi"]["jumlah_indikator"] = 99
    berkas = tmp_path / "ekspektasi-lain.yaml"
    berkas.write_text(yaml.safe_dump(data, allow_unicode=True), encoding="utf-8")

    with pytest.raises(JumlahIndikatorTidakSesuai, match="99.*ditemukan 2"):
        run(workbook_versi_lain, tmp_path / "x.db", tmp_path / "x.md", muat(berkas))
