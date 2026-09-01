"""Validasi dan penyimpanan unggahan realisasi massal milik operator."""

from __future__ import annotations

import io
import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from hashlib import sha256
from pathlib import Path
from uuid import uuid4
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy.orm import Session

from ..models import JenisNilai
from ..repositories import indikator as repo_indikator
from ..repositories import tata_kelola as repo_tata_kelola
from . import Penolakan
from . import bukti as svc_bukti

SHEET_BAKU = "Data Realisasi"
KOLOM_WAJIB = ("id_indikator", "tahun", "sumber")
KOLOM_DIDUKUNG = {
    "wilayah_kode",
    "id_indikator",
    "nama_indikator",
    "tahun",
    "jenis",
    "periode",
    "nilai",
    "nilai_teks",
    "sumber",
    "catatan",
}
MAKS_BARIS = 2_000
MIME_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
ANGKA_BIASA = re.compile(r"^[+-]?\d+(?:[.,]\d+)?$")
ANGKA_RIBUAN_LOKAL = re.compile(r"^[+-]?\d{1,3}(?:\.\d{3})+,\d+$")


@dataclass(frozen=True)
class BarisImpor:
    id_indikator: str
    tahun: int
    periode: int | None
    nilai: float | None
    nilai_teks: str | None
    sumber: str
    catatan: str | None


def _teks(value: object) -> str:
    return "" if value is None else str(value).strip()


def _angka(value: object) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        raise ValueError
    if isinstance(value, (int, float, Decimal)):
        return float(value)
    teks = _teks(value)
    if ANGKA_RIBUAN_LOKAL.fullmatch(teks):
        normal = teks.replace(".", "").replace(",", ".")
    elif ANGKA_BIASA.fullmatch(teks):
        normal = teks.replace(",", ".")
    else:
        return None
    try:
        return float(Decimal(normal))
    except InvalidOperation as exc:
        raise ValueError from exc


def _pilih_sheet(workbook, wilayah_kode: str):
    if wilayah_kode in workbook.sheetnames:
        return workbook[wilayah_kode]
    if SHEET_BAKU in workbook.sheetnames:
        return workbook[SHEET_BAKU]
    raise ValueError(f"Sheet harus bernama '{SHEET_BAKU}' atau '{wilayah_kode}'")


def baca(isi: bytes, wilayah_kode: str) -> list[BarisImpor] | Penolakan:
    """Baca satu sheet wilayah dan kembalikan seluruh baris valid sekaligus."""
    try:
        workbook = load_workbook(io.BytesIO(isi), read_only=True, data_only=True)
        sheet = _pilih_sheet(workbook, wilayah_kode)
    except (BadZipFile, EOFError, InvalidFileException, OSError, ValueError, KeyError) as exc:
        return Penolakan(422, f"Workbook tidak valid: {exc}")

    if sheet.max_row > MAKS_BARIS:
        return Penolakan(422, f"Workbook melebihi batas {MAKS_BARIS - 1} baris data")

    header = {_teks(cell.value).lower(): index for index, cell in enumerate(sheet[1]) if _teks(cell.value)}
    tidak_dikenal = sorted(set(header) - KOLOM_DIDUKUNG)
    hilang = [nama for nama in KOLOM_WAJIB if nama not in header]
    if hilang:
        return Penolakan(422, f"Kolom wajib tidak ditemukan: {', '.join(hilang)}")
    if tidak_dikenal:
        return Penolakan(422, f"Kolom tidak didukung: {', '.join(tidak_dikenal)}")

    hasil: list[BarisImpor] = []
    kunci: set[tuple[str, int, int | None]] = set()
    for nomor, cells in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        def ambil(nama: str, cells=cells) -> object:
            index = header.get(nama)
            return None if index is None or index >= len(cells) else cells[index]

        if not any(value not in (None, "") for value in cells):
            continue
        id_indikator = _teks(ambil("id_indikator")).upper()
        sumber = _teks(ambil("sumber"))
        wilayah = _teks(ambil("wilayah_kode"))
        jenis = _teks(ambil("jenis")).lower() or JenisNilai.REALISASI
        try:
            tahun = int(ambil("tahun"))
            periode_mentah = ambil("periode")
            periode = None if periode_mentah in (None, "") else int(periode_mentah)
            nilai = _angka(ambil("nilai"))
        except (TypeError, ValueError):
            return Penolakan(422, f"Baris {nomor}: tahun, periode, atau nilai angka tidak valid")
        nilai_teks = _teks(ambil("nilai_teks")) or None
        if nilai is None and nilai_teks is None:
            mentah = _teks(ambil("nilai"))
            nilai_teks = mentah or None

        if wilayah and wilayah != wilayah_kode:
            return Penolakan(403, f"Baris {nomor}: wilayah {wilayah} bukan wilayah akun {wilayah_kode}")
        if jenis != JenisNilai.REALISASI:
            return Penolakan(403, f"Baris {nomor}: operator hanya dapat mengunggah realisasi")
        if not id_indikator:
            return Penolakan(422, f"Baris {nomor}: ID indikator wajib diisi")
        if not 2000 <= tahun <= 2045:
            return Penolakan(422, f"Baris {nomor}: tahun harus antara 2000 dan 2045")
        if periode not in (None, 1, 2, 3, 4):
            return Penolakan(422, f"Baris {nomor}: periode harus kosong atau 1–4")
        if nilai is not None and nilai_teks is not None:
            return Penolakan(422, f"Baris {nomor}: isi hanya salah satu dari nilai atau nilai_teks")
        if nilai is None and nilai_teks is None:
            return Penolakan(422, f"Baris {nomor}: nilai wajib diisi")
        if not sumber:
            return Penolakan(422, f"Baris {nomor}: sumber wajib diisi")
        identitas = (id_indikator, tahun, periode)
        if identitas in kunci:
            return Penolakan(422, f"Baris {nomor}: indikator/tahun/periode duplikat")
        kunci.add(identitas)
        hasil.append(
            BarisImpor(
                id_indikator=id_indikator,
                tahun=tahun,
                periode=periode,
                nilai=nilai,
                nilai_teks=nilai_teks,
                sumber=sumber,
                catatan=_teks(ambil("catatan")) or None,
            )
        )
    if not hasil:
        return Penolakan(422, "Workbook tidak memuat baris data")
    return hasil


def simpan(
    session: Session,
    *,
    baris: list[BarisImpor],
    wilayah_kode: str,
    pengusul_id: int,
    nama_file: str,
    isi: bytes,
) -> dict[str, object] | Penolakan:
    """Buat satu usulan per baris; workbook yang sama menjadi bukti batch."""
    id_sah = {item.id_indikator for item in repo_indikator.semua_ringkas(session)}
    tidak_sah = sorted({item.id_indikator for item in baris} - id_sah)
    if tidak_sah:
        return Penolakan(422, f"ID indikator tidak dikenal: {', '.join(tidak_sah[:10])}")

    batch_id = uuid4().hex
    usulan = [
        repo_tata_kelola.buat_usulan(
            session,
            id_indikator=item.id_indikator,
            wilayah_kode=wilayah_kode,
            tahun=item.tahun,
            jenis=JenisNilai.REALISASI,
            periode=item.periode,
            nilai=item.nilai,
            nilai_teks=item.nilai_teks,
            sumber=item.sumber,
            catatan=item.catatan,
            batch_id=batch_id,
            pengusul_id=pengusul_id,
        )
        for item in baris
    ]
    bukti = svc_bukti.simpan(usulan[0].id, nama_file, isi, MIME_XLSX)
    for item in usulan:
        repo_tata_kelola.catat_bukti(
            session,
            usulan_id=item.id,
            nama_file=bukti.nama_file,
            path_file=str(bukti.path_file),
            mime_type=bukti.mime_type,
            ukuran=bukti.ukuran,
            checksum_sha256=bukti.checksum_sha256,
        )
    jumlah_angka = sum(item.nilai is not None for item in baris)
    repo_tata_kelola.catat_aktivitas(
        session,
        pengguna_id=pengusul_id,
        aksi="UNGGAH_USULAN_MASSAL",
        objek_tipe="usulan_nilai",
        objek_id=sha256(isi).hexdigest()[:16],
        detail={"wilayah": wilayah_kode, "jumlah": len(baris), "nama_file": Path(nama_file).name},
    )
    session.commit()
    return {
        "status": "MENUNGGU_VERIFIKASI",
        "jumlah_usulan": len(baris),
        "jumlah_angka": jumlah_angka,
        "jumlah_teks": len(baris) - jumlah_angka,
        "wilayah_kode": wilayah_kode,
        "batch_id": batch_id,
    }
