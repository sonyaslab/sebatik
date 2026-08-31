"""Penyusunan berkas ekspor: CSV, XLSX, PDF, dan paket ZIP."""

from __future__ import annotations

import csv
import zipfile
from collections.abc import Iterable, Sequence
from io import BytesIO, StringIO
from typing import Any

from openpyxl import Workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

from ..models import Indikator

# Label kolom ekspor. Urutan kunci menentukan urutan kolom.
HEADER_EKSPOR: dict[str, str] = {
    "id_indikator": "ID Indikator",
    "nama_indikator": "Nama Indikator",
    "kategori": "Kategori",
    "kelompok": "Kelompok Indikator",
    "satuan": "Satuan",
    "tim_pjk": "Tim PJK",
    "opd_pengampu": "OPD Pengampu",
    "status_metadata": "Status Metadata",
    "tahun_terakhir": "Tahun Terakhir Data",
    "is_proxy": "Menggunakan Proxy",
}

# BOM agar Excel di Windows membuka CSV UTF-8 tanpa mengacak karakter.
BOM = "﻿"


def _sel(indikator: Indikator, field: str) -> Any:
    nilai = getattr(indikator, field)
    if field == "is_proxy":
        return "Ya" if nilai else "Tidak"
    return nilai


def csv_indikator(daftar: Iterable[Indikator]) -> str:
    keluaran = StringIO()
    penulis = csv.DictWriter(keluaran, fieldnames=list(HEADER_EKSPOR.values()))
    penulis.writeheader()
    for indikator in daftar:
        penulis.writerow({label: _sel(indikator, field) for field, label in HEADER_EKSPOR.items()})
    return BOM + keluaran.getvalue()


def xlsx_indikator(daftar: Iterable[Indikator]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Indikator"
    ws.append(list(HEADER_EKSPOR.values()))
    for indikator in daftar:
        ws.append([_sel(indikator, field) for field in HEADER_EKSPOR])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for sel in ws[1]:
        sel.font = sel.font.copy(bold=True)
    aliran = BytesIO()
    wb.save(aliran)
    return aliran.getvalue()


def csv_nilai_indikator(id_indikator: str, nama_indikator: str, nilai: Sequence[Any]) -> str:
    keluaran = StringIO()
    penulis = csv.writer(keluaran)
    penulis.writerow(["ID Indikator", "Nama Indikator", "Tahun", "Jenis", "Nilai", "Sumber"])
    for baris in nilai:
        penulis.writerow([id_indikator, nama_indikator, baris.tahun, baris.jenis, baris.nilai, baris.sumber])
    return BOM + keluaran.getvalue()


def xlsx_katalog_metadata(baris: Sequence[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Katalog Metadata"
    ws.append(
        [
            "ID Indikator",
            "Nama Indikator",
            "Definisi",
            "Rumus Mentah",
            "Interpretasi",
            "Sumber Data",
            "Frekuensi",
            "Sumber Metadata",
        ]
    )
    for item in baris:
        ws.append(list(item.values()))
    aliran = BytesIO()
    wb.save(aliran)
    return aliran.getvalue()


def pdf_katalog_metadata(baris: Sequence[dict[str, Any]]) -> bytes:
    aliran = BytesIO()
    lembar = canvas.Canvas(aliran, pagesize=A4)
    _, tinggi = A4
    y = tinggi - 45
    lembar.setFont("Helvetica-Bold", 15)
    lembar.drawString(40, y, "Katalog Metadata Indikator SEBATIK")
    y -= 28
    for item in baris:
        if y < 70:
            lembar.showPage()
            y = tinggi - 45
        lembar.setFont("Helvetica-Bold", 8)
        lembar.drawString(40, y, f"{item['id_indikator']} - {item['nama_indikator'][:85]}")
        y -= 11
        lembar.setFont("Helvetica", 7)
        teks = (item.get("definisi") or "Metadata belum tersedia")[:350]
        for mulai in range(0, len(teks), 105):
            lembar.drawString(45, y, teks[mulai : mulai + 105])
            y -= 9
        y -= 5
    lembar.save()
    return aliran.getvalue()


def zip_paket(tabel: dict[str, list[dict[str, Any]]], katalog: Sequence[dict[str, Any]]) -> bytes:
    aliran = BytesIO()
    with zipfile.ZipFile(aliran, "w", zipfile.ZIP_DEFLATED) as arsip:
        for nama, baris in tabel.items():
            keluaran = StringIO()
            penulis = csv.DictWriter(keluaran, fieldnames=list(baris[0]) if baris else [])
            penulis.writeheader()
            penulis.writerows(baris)
            arsip.writestr(f"{nama}.csv", BOM + keluaran.getvalue())
        arsip.writestr("katalog-metadata.xlsx", xlsx_katalog_metadata(katalog))
        arsip.writestr("katalog-metadata.pdf", pdf_katalog_metadata(katalog))
    return aliran.getvalue()


# ---------------------------------------------------------------------------
# Perakitan paket unduhan.
#
# Fungsi di atas hanya membentuk berkas dari baris yang sudah jadi. Fungsi di
# bawah membaca repository dan menyusun barisnya — tetap di luar router karena
# isinya pemilihan kolom, bukan HTTP (backend.md §1.2).
# ---------------------------------------------------------------------------

from sqlalchemy.orm import Session  # noqa: E402

from ..models import KODE_PROVINSI  # noqa: E402
from ..repositories import indikator as repo_indikator  # noqa: E402
from ..repositories import nilai as repo_nilai  # noqa: E402


def paket_lengkap(session: Session) -> bytes:
    """ZIP berisi tabel indikator, nilai, metadata, dan katalognya."""
    daftar = repo_indikator.daftar_ekspor(session)
    katalog = []
    tabel_indikator = []
    for item in daftar:
        metadata = repo_indikator.ambil_metadata(session, item.id_indikator)
        katalog.append(
            {
                "id_indikator": item.id_indikator,
                "nama_indikator": item.nama_indikator,
                "definisi": metadata.definisi if metadata else None,
                "rumus_mentah": metadata.rumus_mentah if metadata else None,
                "interpretasi": metadata.interpretasi if metadata else None,
                "sumber_data": metadata.sumber_data if metadata else None,
                "frekuensi": metadata.frekuensi if metadata else None,
                "sumber_metadata": metadata.sumber_metadata if metadata else None,
            }
        )
        tabel_indikator.append(
            {
                "id_indikator": item.id_indikator,
                "kategori": item.kategori,
                "nomor": item.nomor,
                "kode_indikator": item.kode_indikator,
                "nama_indikator": item.nama_indikator,
                "kelompok": item.kelompok,
                "satuan": item.satuan,
                "opd_pengampu": item.opd_pengampu,
                "tim_pjk": item.tim_pjk,
                "status_metadata": item.status_metadata,
                "tahun_terakhir": item.tahun_terakhir,
                "is_proxy": item.is_proxy,
                "arah_baik": item.arah_baik,
            }
        )

    tabel_nilai = [
        {
            "id_indikator": baris.id_indikator,
            "wilayah_kode": baris.wilayah_kode,
            "tahun": baris.tahun,
            "jenis": baris.jenis,
            "periode": baris.periode,
            "nilai": baris.nilai,
            "nilai_teks": baris.nilai_teks,
            "sumber": baris.sumber,
        }
        for item in daftar
        for baris in repo_nilai.seri_lengkap(session, item.id_indikator, KODE_PROVINSI)
    ]
    # Katalog memakai nama indikator sebagai judul bagian; tabel metadata tidak
    # perlu mengulangnya karena sudah ada di tabel indikator.
    tabel_metadata = [{k: v for k, v in baris.items() if k != "nama_indikator"} for baris in katalog]

    return zip_paket(
        {
            "indikator": tabel_indikator,
            "nilai_indikator": tabel_nilai,
            "metadata_indikator": tabel_metadata,
        },
        katalog,
    )


def csv_semua_indikator(session: Session) -> str:
    return csv_indikator(repo_indikator.daftar_ekspor(session))


def xlsx_semua_indikator(session: Session) -> bytes:
    return xlsx_indikator(repo_indikator.daftar_ekspor(session))


def csv_satu_indikator(session: Session, indikator: Indikator) -> str:
    return csv_nilai_indikator(
        indikator.id_indikator,
        indikator.nama_indikator,
        repo_nilai.seri(session, indikator.id_indikator, KODE_PROVINSI),
    )
