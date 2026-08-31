"""Perbarui sumber data indikator dari workbook pemetaan.

Workbook disusun dalam urutan master: 10 ISV lalu 76 IUP. Nama tetap
divalidasi sebelum transaksi agar satu baris yang bergeser tidak memasangkan
sumber ke indikator yang salah.
"""

from __future__ import annotations

import argparse
import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select

from backend.app.database import SessionLocal
from backend.app.models import Indikator, MetadataIndikator


def _normalisasi(teks: str | None) -> str:
    return re.sub(r"[^a-z0-9]+", " ", (teks or "").casefold()).strip()


def _nama_selaras(nama_workbook: str, nama_basis_data: str) -> bool:
    kiri, kanan = _normalisasi(nama_workbook), _normalisasi(nama_basis_data)
    return (
        kiri == kanan or (len(kiri) >= 30 and kanan.startswith(kiri)) or (len(kanan) >= 30 and kiri.startswith(kanan))
    )


def baca_workbook(path: Path) -> list[tuple[str, str]]:
    buku = load_workbook(path, read_only=True, data_only=True)
    lembar = buku["Pemetaan Sumber Data"]
    hasil: list[tuple[str, str]] = []
    for nama, sumber in lembar.iter_rows(min_row=2, max_col=2, values_only=True):
        if nama and sumber:
            hasil.append((str(nama).strip(), str(sumber).strip()))
    return hasil


def _pasangkan(indikator: list[Indikator], sumber: list[tuple[str, str]]) -> list[tuple[Indikator, str]]:
    tersisa = list(indikator)
    hasil: list[tuple[Indikator, str]] = []
    gagal: list[str] = []
    for nama, nilai_sumber in sumber:
        kandidat = [item for item in tersisa if _normalisasi(item.nama_indikator) == _normalisasi(nama)]
        if not kandidat:
            kandidat = [item for item in tersisa if _nama_selaras(nama, item.nama_indikator)]
        if not kandidat:
            gagal.append(nama)
            continue
        item = kandidat[0]
        tersisa.remove(item)
        hasil.append((item, nilai_sumber))
    if gagal or tersisa:
        raise ValueError(
            "Nama indikator belum dapat dipasangkan. "
            f"Workbook: {gagal[:8]}; basis data: {[x.nama_indikator for x in tersisa[:8]]}"
        )
    return hasil


def jalankan(workbook_path: Path, backup_dir: Path, periksa: bool = False) -> tuple[int, Path | None]:
    sumber = baca_workbook(workbook_path)
    with SessionLocal() as session:
        indikator = list(session.scalars(select(Indikator).order_by(Indikator.kategori.asc(), Indikator.nomor)))
        if len(sumber) != len(indikator):
            raise ValueError(f"Jumlah baris berbeda: workbook {len(sumber)}, basis data {len(indikator)}.")

        pasangan = _pasangkan(indikator, sumber)

        if periksa:
            return len(indikator), None

        backup_dir.mkdir(parents=True, exist_ok=True)
        backup_path = backup_dir / f"sumber-data-{datetime.now():%Y%m%d-%H%M%S}.json"
        backup_path.write_text(
            json.dumps(
                [
                    {
                        "id_indikator": item.id_indikator,
                        "nama_indikator": item.nama_indikator,
                        "sumber_data_indikator": item.sumber_data,
                        "sumber_data_metadata": (
                            session.get(MetadataIndikator, item.id_indikator).sumber_data
                            if session.get(MetadataIndikator, item.id_indikator)
                            else None
                        ),
                    }
                    for item in indikator
                ],
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )

        for item, nilai_sumber in pasangan:
            item.sumber_data = nilai_sumber
            metadata = session.get(MetadataIndikator, item.id_indikator)
            if metadata:
                metadata.sumber_data = nilai_sumber
        session.commit()
        return len(indikator), backup_path


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--periksa", action="store_true", help="Validasi tanpa menulis basis data")
    parser.add_argument("--backup-dir", type=Path, default=Path("data/processed/cadangan"))
    args = parser.parse_args()
    jumlah, backup = jalankan(args.workbook, args.backup_dir, args.periksa)
    print(f"Sumber data tervalidasi: {jumlah} indikator")
    if backup:
        print(f"Cadangan nilai lama: {backup}")
