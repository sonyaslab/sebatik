"""Ekspor JSON fixture seed indikator dari Excel klasifikasi ISV/IUP.

Jalankan sekali secara manual setelah
`data/raw/BASIS_DATA_INDIKATOR_ISV-IUP_KALTARA.xlsx` tersedia. Hasilnya
(`backend/app/data/indikator_seed.json`) di-commit ke git; skrip ini sendiri
TIDAK dijalankan otomatis saat deploy — lihat `backend/app/cli.py`
(`seed-indikator`) untuk pemakaian fixture ini saat runtime.

Kolom "Arah Pembangunan" di sheet sumber berarti dua hal berbeda tergantung
kategori: untuk ISV ia kalimat arah pembangunan (`indikator.arah_pembangunan`),
untuk IUP ia kode pilar Indonesia Emas (`indikator.arah_ie`). Jangan disatukan.

Format `id_indikator` di sini tiga digit (`ISV-001`) dan dibaca apa adanya dari
kolomnya. `src/etl/common.py::indicator_id()` menghasilkan format dua digit
untuk workbook lama yang bentuknya berbeda — jangan dipakai di sini.

Pemakaian::

    python scripts/ekspor_seed_indikator.py \
        data/raw/BASIS_DATA_INDIKATOR_ISV-IUP_KALTARA.xlsx \
        backend/app/data/indikator_seed.json
"""

from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

# Membuat `src` dapat diimpor ketika skrip dijalankan langsung sebagai
# `python scripts/ekspor_seed_indikator.py` (bukan lewat pytest, yang sudah
# menaruh akar repo di sys.path lewat konfigurasi `pythonpath` di pyproject.toml).
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from src.etl.common import clean_text, parse_angka  # noqa: E402
from src.etl.transform.proxy import ekstrak_proxy  # noqa: E402

SHEET_INDIKATOR = "Basis Data Indikator"
SHEET_NILAI = "Data Target-Realisasi"
JUMLAH_INDIKATOR_DIHARAPKAN = 86

# Tiga kolom catatan mirip di sheet sumber, digabung jadi satu
# `indikator.catatan_teknis` supaya tidak ada informasi yang hilang.
KOLOM_CATATAN = (
    "Catatan Kualitas Data",
    "Keterangan (Rakor Kaltara)",
    "Keterangan RPJMD / Catatan Kaltara",
)


def _header(ws: Any) -> dict[str, int]:
    """Label header (baris 1) -> nomor kolom, sepola dengan src/etl/extract/master.py."""
    hasil: dict[str, int] = {}
    for kolom in range(1, ws.max_column + 1):
        label = clean_text(ws.cell(1, kolom).value)
        if label:
            hasil[label] = kolom
    return hasil


def baca_indikator_dan_metadata(wb: Any) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Baris `indikator` + `metadata_indikator` dari sheet `Basis Data Indikator`."""
    ws = wb[SHEET_INDIKATOR]
    header = _header(ws)

    def sel(baris: int, label: str) -> Any:
        kolom = header.get(label)
        return ws.cell(baris, kolom).value if kolom else None

    indikator: list[dict[str, Any]] = []
    metadata: list[dict[str, Any]] = []

    for baris in range(2, ws.max_row + 1):
        id_indikator = clean_text(sel(baris, "ID Indikator"))
        if not id_indikator:
            continue

        kategori = (clean_text(sel(baris, "Kategori")) or "").upper()
        suffix = id_indikator.split("-")[-1]
        nomor = int(suffix) if suffix.isdigit() else None

        # "Arah Pembangunan" berarti dua hal berbeda tergantung kategori —
        # lihat catatan di docstring modul ini.
        kolom_arah = clean_text(sel(baris, "Arah Pembangunan"))
        arah_pembangunan = kolom_arah if kategori == "ISV" else None
        arah_ie = kolom_arah if kategori == "IUP" else None

        is_proxy, nama_proxy = ekstrak_proxy(
            sel(baris, "Indikator Proxy?"),
            sel(baris, "Keterangan RPJMD / Catatan Kaltara"),
        )

        catatan_gabungan = "\n".join(
            f"[{label}] {teks}" for label in KOLOM_CATATAN if (teks := clean_text(sel(baris, label)))
        )

        sumber_data = clean_text(sel(baris, "Sumber Data (RPJPD Provinsi)"))
        frekuensi = clean_text(sel(baris, "Frekuensi (RPJPD Provinsi)"))
        status_metadata = clean_text(sel(baris, "Status Metadata"))
        tahun_terakhir = parse_angka(sel(baris, "Tahun Data Terakhir"))

        indikator.append(
            {
                "id_indikator": id_indikator,
                "kategori": kategori,
                "nomor": nomor,
                "kode_indikator": clean_text(sel(baris, "Kode Indikator")),
                "nama_indikator": clean_text(sel(baris, "Nama Indikator (RPJPD Provinsi / dipakai Kaltara)")),
                "kelompok": clean_text(sel(baris, "Kelompok / Pilar")),
                "arah_pembangunan": arah_pembangunan,
                "arah_ie": arah_ie,
                "opd_pengampu": clean_text(sel(baris, "Perangkat Daerah Pengampu (Kaltara)")),
                "sumber_data": sumber_data,
                "frekuensi": frekuensi,
                "status_ketersediaan": clean_text(sel(baris, "Ketersediaan Data")),
                "status_metadata": status_metadata,
                "periode_data": clean_text(sel(baris, "Periode Data")),
                "tahun_terakhir": int(tahun_terakhir) if tahun_terakhir is not None else None,
                "is_proxy": bool(is_proxy),
                "nama_proxy": nama_proxy,
                "catatan_teknis": catatan_gabungan or None,
            }
        )
        metadata.append(
            {
                "id_indikator": id_indikator,
                "definisi": clean_text(sel(baris, "Definisi (RPJPD Provinsi)")),
                "rumus_mentah": clean_text(sel(baris, "Rumus Perhitungan (RPJPD Provinsi)")),
                "interpretasi": clean_text(sel(baris, "Interpretasi (RPJPD Provinsi)")),
                "sumber_data": sumber_data,
                "frekuensi": frekuensi,
                "status_metadata": status_metadata,
            }
        )

    return indikator, metadata


def baca_nilai(wb: Any) -> list[dict[str, Any]]:
    """Baris `nilai_indikator` (provinsi, tahunan) dari sheet `Data Target-Realisasi`."""
    ws = wb[SHEET_NILAI]
    header = _header(ws)

    def sel(baris: int, label: str) -> Any:
        kolom = header.get(label)
        return ws.cell(baris, kolom).value if kolom else None

    hasil: list[dict[str, Any]] = []
    for baris in range(2, ws.max_row + 1):
        id_indikator = clean_text(sel(baris, "ID Indikator"))
        if not id_indikator:
            continue

        jenis_teks = (clean_text(sel(baris, "Jenis Nilai")) or "").casefold()
        if jenis_teks == "realisasi":
            jenis = "realisasi"
        elif jenis_teks == "target":
            jenis = "target"
        else:
            continue

        tahun = parse_angka(sel(baris, "Tahun"))
        if tahun is None:
            continue

        hasil.append(
            {
                "id_indikator": id_indikator,
                "wilayah_kode": "65",
                "tahun": int(tahun),
                "jenis": jenis,
                "periode": None,
                "nilai": parse_angka(sel(baris, "Nilai (Angka)")),
                "nilai_teks": clean_text(sel(baris, "Nilai (Teks Asli)")),
                "satuan_catatan": clean_text(sel(baris, "Satuan/Catatan")),
                "sumber": "seed_awal:BASIS_DATA_INDIKATOR_ISV-IUP_KALTARA.xlsx",
            }
        )
    return hasil


def main(sumber: Path, target: Path) -> None:
    wb = load_workbook(sumber, data_only=True, read_only=True)
    indikator, metadata = baca_indikator_dan_metadata(wb)
    nilai = baca_nilai(wb)

    if len(indikator) != JUMLAH_INDIKATOR_DIHARAPKAN:
        raise SystemExit(
            f"Diharapkan {JUMLAH_INDIKATOR_DIHARAPKAN} baris indikator, didapat {len(indikator)}. "
            "Periksa apakah sheet 'Basis Data Indikator' berubah struktur."
        )

    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(
        json.dumps(
            {"indikator": indikator, "metadata_indikator": metadata, "nilai_indikator": nilai},
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    print(f"Ditulis {len(indikator)} indikator, {len(nilai)} nilai -> {target}")


if __name__ == "__main__":
    main(Path(sys.argv[1]), Path(sys.argv[2]))
