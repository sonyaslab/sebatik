"""Adaptor .xlsx -> bentuk sumber dataset database.

Satu-satunya tempat workbook klasifikasi ISV/IUP dibaca menjadi struktur
`{"source": ..., "sheets": {...}}`. Keluarannya langsung dapat diumpankan ke
`src.etl.database.transformasi_sumber_database()`, sehingga tidak ada logika
pemetaan kolom baru di sini — modul ini murni membaca sel.

Dipakai dua pemanggil: gerbang API (`backend/app/services/unggahan.py`) dan
CLI (`scripts/kelola_database.py excel`), supaya keduanya memakai konverter
yang sama persis. `tools/import_classified_workbook.py` juga memanggilnya
agar tidak ada dua salinan logika baca.
"""

from __future__ import annotations

import zipfile
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .database import DatasetTidakValid

SHEET_WAJIB = ("Basis Data Indikator", "Data Target-Realisasi")


def _nilai_sel(sel: Any) -> Any:
    """Sel bertipe kaya (tanggal, waktu) diratakan jadi teks.

    Sepola dengan `tools/import_classified_workbook.py` supaya dataset hasil
    jalur Excel identik dengan dataset jalur JSON lama.
    """
    mentah = sel.value
    if mentah is None or isinstance(mentah, (str, int, float, bool)):
        return mentah
    return str(mentah)


def _baca_sheet(sheet: Any) -> list[list[Any]]:
    """Baca sampai baris data terakhir, bukan sampai batas format Excel.

    Beberapa editor menandai seluruh kolom sebagai used range. Menghentikan
    pembacaan pada baris kosong pertama setelah data menjaga unggahan tetap
    cepat tanpa mengubah kontrak bahwa baris master/nilai harus rapat.
    """
    hasil: list[list[Any]] = []
    for baris in sheet.iter_rows():
        nilai = [_nilai_sel(sel) for sel in baris]
        if hasil and all(item is None for item in nilai):
            break
        hasil.append(nilai)
    return hasil


def baca_workbook(isi: bytes, nama_berkas: str) -> dict[str, Any]:
    """Ubah byte .xlsx menjadi bentuk `{"source": ..., "sheets": {...}}`.

    `data_only=True` wajib: sebagian sel di berkas nyata berisi rumus, dan
    tanpa flag ini yang terbaca adalah teks rumusnya, bukan hasilnya.
    """
    try:
        workbook = load_workbook(BytesIO(isi), data_only=True, read_only=True)
    except (InvalidFileException, zipfile.BadZipFile, OSError, ValueError) as exc:
        raise DatasetTidakValid(f"Berkas '{nama_berkas}' bukan berkas .xlsx yang dapat dibaca: {exc}") from exc

    try:
        tersedia = set(workbook.sheetnames)
        hilang = [nama for nama in SHEET_WAJIB if nama not in tersedia]
        if hilang:
            raise DatasetTidakValid(f"Sheet '{hilang[0]}' tidak ditemukan di workbook")
        sheets = {nama: _baca_sheet(workbook[nama]) for nama in SHEET_WAJIB}
    finally:
        # Mode read_only menahan handle berkas sampai ditutup eksplisit.
        workbook.close()

    return {"source": nama_berkas, "sheets": sheets}
