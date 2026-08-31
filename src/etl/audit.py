"""Audit read-only workbook ISV-IUP dan tulis laporan Markdown."""

from __future__ import annotations

import argparse
import re
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

from .common import clean_text, indicator_id, parse_angka
from .config import bawaan
from .transform import kategori_dari_nomor, nomor_dalam_kategori

# Audit dan pipeline membaca konfigurasi yang sama supaya keduanya tidak
# menyimpang saat versi workbook berganti (etl.md §7).
KONFIGURASI = bawaan()
SHEETS = list(KONFIGURASI.sheet_audit)
TAHUN_VALID = range(*KONFIGURASI.tahun_valid)


def last_used(ws):
    rows = [r for r in range(1, ws.max_row + 1) if any(ws.cell(r, c).value is not None for c in range(1, ws.max_column + 1))]
    cols = [c for c in range(1, ws.max_column + 1) if any(ws.cell(r, c).value is not None for r in range(1, ws.max_row + 1))]
    return (max(rows, default=0), max(cols, default=0))


def detect_header_rows(ws):
    first_values = {clean_text(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)}
    if "Indikator" in first_values or "No" in first_values:
        years = sum(
            parse_angka(ws.cell(2, c).value) in TAHUN_VALID
            if parse_angka(ws.cell(2, c).value) is not None else False
            for c in range(1, ws.max_column + 1)
        )
        return [1, 2] if years >= 3 else [1]
    best = (0, 1)
    for r in range(1, min(10, ws.max_row) + 1):
        vals = [clean_text(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        score = sum(v is not None for v in vals) + 4 * sum(v in {"Kategori", "No", "Indikator"} for v in vals)
        if score > best[0]:
            best = (score, r)
    start = best[1]
    second = start + 1
    has_years = second <= ws.max_row and sum(parse_angka(ws.cell(second, c).value) in TAHUN_VALID if parse_angka(ws.cell(second, c).value) is not None else False for c in range(1, ws.max_column + 1)) >= 3
    return [start, second] if has_years else [start]


def combined_headers(ws, rows, ncols):
    out = []
    for c in range(1, ncols + 1):
        parts = []
        for r in rows:
            value = clean_text(ws.cell(r, c).value)
            if value:
                parts.append(value)
        out.append(" | ".join(parts) or f"kolom_{c}")
    return out


def logical_type(values):
    types = Counter()
    for v in values:
        if v is None or str(v).strip() == "":
            continue
        if isinstance(v, bool): types["boolean"] += 1
        elif isinstance(v, (int, float)): types["angka"] += 1
        elif getattr(v, "year", None): types["tanggal"] += 1
        elif isinstance(v, str): types["teks"] += 1
        else: types[type(v).__name__] += 1
    return ", ".join(f"{k}:{v}" for k, v in types.items()) or "kosong"


def daftar_id_diharapkan() -> list[str]:
    """Seluruh ID indikator yang seharusnya ada, diturunkan dari konfigurasi.

    Jumlah ISV berasal dari batas penomoran menyambung; sisanya IUP hingga
    total yang diharapkan. Tidak ada angka 86 atau 76 yang dikodekan di sini.
    """
    isv = KONFIGURASI.isv_nomor_maksimum
    iup = KONFIGURASI.jumlah_indikator_diharapkan - isv
    return [f"ISV-{i:02d}" for i in range(1, isv + 1)] + [f"IUP-{i:02d}" for i in range(1, iup + 1)]


def extract_indicators(ws, sheet):
    rows = []
    isv_maks = KONFIGURASI.isv_nomor_maksimum
    if sheet == KONFIGURASI.master.sheet:
        sequence = dict.fromkeys(KONFIGURASI.kategori, 0)
        for r in KONFIGURASI.master.baris.rentang(ws.max_row):
            cat, num, name = ws.cell(r, 1).value, ws.cell(r, 4).value, ws.cell(r, 5).value
            cat = (clean_text(cat) or "").upper()
            if cat not in sequence:
                continue
            sequence[cat] += 1
            iid = indicator_id(cat, sequence[cat])
            if iid: rows.append((iid, clean_text(name), r))
        return rows
    sumber = next((x for x in KONFIGURASI.nilai if x.sheet == sheet), None)
    if sumber is not None and sumber.identitas == "kategori_nomor":
        cat = num = name = None
        for r in sumber.baris.rentang(ws.max_row):
            cat = ws.cell(r, sumber.kolom_kategori).value or cat
            num = ws.cell(r, sumber.kolom_nomor).value or num
            name = ws.cell(r, 3).value or name
            if sumber.kolom_jenis and clean_text(ws.cell(r, sumber.kolom_jenis).value) == "Target":
                kategori = (clean_text(cat) or "").upper()
                iid = indicator_id(kategori, nomor_dalam_kategori(num, kategori, isv_maks))
                if iid: rows.append((iid, clean_text(name), r))
        return rows
    # Sheet lama memakai nomor global menyambung; IUP melanjutkan penomoran ISV.
    baris = sumber.baris if sumber is not None else KONFIGURASI.pemilik.baris
    for r in baris.rentang(min(last_used(ws)[0], ws.max_row)):
        num, name = ws.cell(r, 1).value, ws.cell(r, 2).value
        n = parse_angka(num)
        if n is None or not clean_text(name): continue
        category = kategori_dari_nomor(n, isv_maks)
        iid = indicator_id(category, nomor_dalam_kategori(n, category, isv_maks))
        if iid: rows.append((iid, clean_text(name), r))
    return rows


def run(workbook_path: Path, output_path: Path):
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    report = ["# Audit Sumber Data ISV-IUP", "", f"Sumber: `{workbook_path.name}`. Audit bersifat read-only; sel kosong tidak ditafsirkan sebagai nol.", ""]
    mappings = defaultdict(dict)
    numeric_text = []
    whitespace = []
    decimal_styles = defaultdict(set)
    for sheet in SHEETS:
        ws = wb[sheet]
        nrows, ncols = last_used(ws)
        headers_r = detect_header_rows(ws)
        headers = combined_headers(ws, headers_r, ncols)
        merged = [str(x) for x in ws.merged_cells.ranges]
        report += [f"## Sheet: {sheet}", "", f"- Area terpakai: **{nrows} baris x {ncols} kolom** (dimensi tersimpan Excel: {ws.max_row} x {ws.max_column}).", f"- Baris header sebenarnya: **{', '.join(map(str, headers_r))}**.", f"- Merged cell: {len(merged)} rentang" + (f"; contoh `{', '.join(merged[:12])}`." if merged else "."), "", "| Kolom | Tipe data (jumlah sel) | Sel kosong |", "|---|---:|---:|"]
        data_start = max(headers_r) + 1
        for c, header in enumerate(headers, 1):
            values = [ws.cell(r, c).value for r in range(data_start, nrows + 1)]
            blanks = sum(v is None or str(v).strip() == "" for v in values)
            report.append(f"| {header.replace('|','/')} | {logical_type(values)} | {blanks} |")
            for r, v in enumerate(values, data_start):
                if isinstance(v, str) and re.fullmatch(r"\s*[+-]?[\d.,]+\s*", v) and parse_angka(v) is not None:
                    numeric_text.append((sheet, f"{ws.cell(r,c).coordinate}", v))
                    if "," in v: decimal_styles[(sheet, header)].add("koma")
                    if "." in v: decimal_styles[(sheet, header)].add("titik")
        inds = extract_indicators(ws, sheet)
        for iid, name, row in inds:
            mappings[iid][sheet] = (name, row)
            raw = ws.cell(row, 5 if sheet == "form provinsi" else (3 if sheet.startswith("Rakor") else 2)).value
            if isinstance(raw, str) and ("\n" in raw or "  " in raw):
                whitespace.append((sheet, row, raw, clean_text(raw)))
        report.append("")

    mixed = [(s,h,sorted(x)) for (s,h),x in decimal_styles.items() if len(x)>1]
    report += ["## Anomali", "", f"### Angka disimpan sebagai teks ({len(numeric_text)} sel)", "", "| Sheet | Sel | Nilai mentah |", "|---|---|---:|"]
    report += [f"| {s} | {cell} | `{v}` |" for s,cell,v in numeric_text]
    report += ["", f"### Campuran pemisah desimal ({len(mixed)} kolom)", "", "| Sheet | Kolom | Gaya yang ditemukan |", "|---|---|---|"]
    report += [f"| {s} | {h.replace('|','/')} | {', '.join(styles)} |" for s,h,styles in mixed]
    report += ["", f"### Spasi ganda/newline dalam nama indikator ({len(whitespace)} kasus)", "", "| Sheet | Baris | Nama mentah | Hasil normalisasi (untuk pencocokan saja) |", "|---|---:|---|---|"]
    report += [f"| {s} | {r} | {str(raw).replace(chr(10),'<br>').replace('|','/')} | {norm.replace('|','/')} |" for s,r,raw,norm in whitespace]

    name_diffs = []
    for iid, by_sheet in sorted(mappings.items()):
        variants = {clean_text(v[0]).casefold() for v in by_sheet.values() if v[0]}
        if len(variants) > 1:
            name_diffs.append((iid, by_sheet))
    report += ["", f"### Nama berbeda antar-sheet ({len(name_diffs)} indikator)", "", "| ID | Variasi nama per sheet |", "|---|---|"]
    for iid, by_sheet in name_diffs:
        desc = "; ".join(f"{s}: {v[0]}" for s,v in by_sheet.items())
        report.append(f"| {iid} | {desc.replace('|','/')} |")

    report += ["", "## Pemetaan indikator antar-sheet", "", "Kunci pemetaan adalah `Kategori + No`. Pada sheet tanpa kolom Kategori, kategori diinferensikan saat nomor kembali ke 1.", "", "| ID | " + " | ".join(SHEETS) + " |", "|---|" + "---|" * len(SHEETS)]
    all_ids = daftar_id_diharapkan()
    for iid in all_ids:
        cells = []
        for sheet in SHEETS:
            item = mappings.get(iid,{}).get(sheet)
            cells.append(f"baris {item[1]}" if item else "**HILANG**")
        report.append(f"| {iid} | " + " | ".join(cells) + " |")
    report += ["", "### Ringkasan indikator hilang", ""]
    for sheet in SHEETS:
        missing = [iid for iid in all_ids if sheet not in mappings.get(iid,{})]
        report.append(f"- **{sheet}: {len(missing)}** — {', '.join(missing) if missing else 'tidak ada' }.")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text("\n".join(report) + "\n", encoding="utf-8")
    print(f"Audit selesai: {output_path} ({len(numeric_text)} angka-teks, {len(name_diffs)} beda nama)")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--output", type=Path, default=Path("docs/01-audit-data.md"))
    args = parser.parse_args()
    run(args.workbook, args.output)
